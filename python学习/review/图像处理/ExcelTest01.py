# python learing
# author:TNT
import datetime

from openpyxl import Workbook
'''
Python的openpyxl模块让我们可以在Python程序中读取和修改Excel电子表格，
由于微软从Office 2007开始使用了新的文件格式，这使得Office Excel和LibreOffice Calc、OpenOffice Calc是完全兼容的，
这就意味着openpyxl模块也能处理来自这些软件生成的电子表格。

'''

wb = Workbook()
ws = wb.active


ws['A1'] = 42
ws.append([1, 2, 3])
ws['A1'] = datetime.datetime.now()
wb.save("sample.xlsx")